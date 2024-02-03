# this is a simple program that send emails for the members that have not yet paid the monthly fee
import openpyxl, smtplib
from pathlib import Path

# connect with ExcelFile
file = openpyxl.load_workbook(Path.home() / Path("OneDrive", "Escritorio", "members.xlsx"))

sheets = file.sheetnames
sheet = file['Sheet1']

lastcol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastcol).value
print(latestMonth)

# find unpaid members
unpaidMembers = {}
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastcol).value
    if payment != 'paid':  # != means not equals
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

# send email
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.starttls()
sender_email = "triumphbcn@gmail.com"
password = input(str("Please enter your password: "))
smtpObj.login(sender_email, password)

for name, email in unpaidMembers.items():
    body = """Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not paid dues 
    for %s. Please make this payment as soon as possible. Thank you!""" %(latestMonth, name, latestMonth)

    print('Sending email to %s...' % email)
    sendMailStatus = smtpObj.sendmail(sender_email, email, body)

    if sendMailStatus != {}:
        print('There was an error sending email to %s: %s' % (email, sendMailStatus))

smtpObj.quit()
